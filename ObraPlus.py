import customtkinter
import openpyxl
from tkinter import ttk
import re
import os
from tkinter import messagebox
from tkinter.constants import LEFT
from datetime import datetime, timedelta
from PIL import Image, ImageTk, ImageDraw 
from tkinter import PhotoImage
import customtkinter
import sqlite3
import tkinter as tk
from tkinter import *
from PIL import Image, ImageTk
from urllib.request import urlretrieve
import pandas as pd
from openpyxl import Workbook
from datetime import datetime, timedelta
from tkinter import filedialog
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle
from reportlab.lib import colors
from reportlab.lib.units import inch
from reportlab.pdfgen import canvas


class TelaLogin:
    def __init__(self):
                # Janela
        self.conn = sqlite3.connect("login.db")
        self.cursor = self.conn.cursor()
        # Crie a tabela de login se ela não existir
        self.cursor.execute('''CREATE TABLE IF NOT EXISTS usuarios (
                            id INTEGER PRIMARY KEY AUTOINCREMENT,
                            usuario TEXT NOT NULL,
                            senha TEXT NOT NULL
                        )''')
        self.conn.commit()

        # verificar se ja esta salvo o cadastro:
        self.cursor.execute("SELECT COUNT(*) FROM usuarios")
        self.numero_registros = self.cursor.fetchone()[0]

        customtkinter.set_appearance_mode("Dark")
        self.janela = customtkinter.CTk()
        self.janela.maxsize(500, 300)
        self.janela.resizable(False, False)
        self.janela.title("ObraPlus")
        self.canvas_etapas = None
        self.canvas_cliente = None
        self.canvas_funcionarios = None
        self.canvas_criado = None
        self.botao_clicado = False
        self.tela_atual = None
        self.etapas_data = []




        # Obtém as dimensões da tela

        self.janela.iconbitmap('Obra_1_.ico')

        largura_tela = self.janela.winfo_screenwidth()
        altura_tela = self.janela.winfo_screenheight()

        # Calcula a posição central da janela
        largura_janela = 500
        altura_janela = 300
        x = (largura_tela - largura_janela) // 2
        y = (altura_tela - altura_janela) // 2

        # Define a geometria da janela para centralizá-la
        self.janela.geometry(f"{largura_janela}x{altura_janela}+{x}+{y}")

        self.canvas_etapas = None
        self.canvas_cliente = None
        self.canvas_funcionarios = None
        self.canvas_criado = None
        self.botao_clicado = False
        self.tela_atual = None




        # criar um frame a direita
        self.frame_login = customtkinter.CTkFrame(
            self.janela, width=500)
        self.frame_login.pack(side=tk.LEFT)
        # Carregue uma imagem (substitua 'sua_imagem.png' pelo caminho da sua imagem)
        self.image_login = customtkinter.CTkImage(
            light_image=Image.open(r"Obra_5_-removebg-preview.png"), size=(300, 300))

        # Crie um widget de imagem e adicione-o ao frame
        self.image_widget_login = customtkinter.CTkLabel(
            self.frame_login, image=self.image_login, width=200, height=1000, text="")
        self.image_widget_login.pack()

        # Edição de elementos na self.janela
        self.texto = customtkinter.CTkLabel(
            self.janela, text="BEM-VINDO", text_color="#fff", font=("Century Gothic Bold", 20))
        self.texto.pack(padx=10, pady=5)

        self.Usuario = customtkinter.CTkEntry(
            self.janela, placeholder_text="Seu Usuario", font=("Century Gothic Bold", 15))
        self.Usuario.pack(padx=10, pady=5)

        self.senha = customtkinter.CTkEntry(self.janela, placeholder_text="Sua Senha", font=(
            "Century Gothic Bold", 15), show="*")
        self.senha.pack(padx=10, pady=5)

        self.botao_login = customtkinter.CTkButton(self.janela, text="Login", font=(
            "Century Gothic Bold", 18), command=self.verificar_login,fg_color="#9932CC",hover_color="#4B0082")
        self.botao_login.pack(padx=10, pady=5)
        self.senha.bind("<Return>", self.verificar_login)

        if self.numero_registros == 0:
            self.botao_cadastro = customtkinter.CTkButton(self.janela, text="Cadastrar", font=(
                "Century Gothic Bold", 18), command=self.cadastrar, width=10,fg_color="#BA55D3",hover_color="#4B0082")
            self.botao_cadastro.pack(padx=10, pady=10)

        # Conecte-se ao banco de dados (ou crie um novo se não existir)
        self.conn = sqlite3.connect("login.db")
        self.cursor = self.conn.cursor()

        self.janela.mainloop()

    def verificar_login(self, event=None):
        self.usuario_cadastrado = self.Usuario.get()
        self.senha_digitada = self.senha.get()

        # Consulte o banco de dados para verificar o login
        self.cursor.execute("SELECT usuario, senha FROM usuarios LIMIT 1")
        self.primeiro_registro = self.cursor.fetchone()

        if self.primeiro_registro is not None:
            usuario_db, senha_db = self.primeiro_registro

            # Verifique se o usuário e senha digitados correspondem ao primeiro registro
            if self.usuario_cadastrado == usuario_db and self.senha_digitada == senha_db:
                # Login bem-sucedido, abra o menu (implemente essa função)
                self.abrir_menu()
            else:
                erro_msg = customtkinter.CTkLabel(self.janela, text=(
                    "ERRO\nUsuário ou senha incorretos!!!"), text_color="red", font=("Century Gothic Bold", 11))
                erro_msg.pack(padx=10, pady=5)
                erro_msg.after(4000, erro_msg.destroy)
        else:
            erro_msg = customtkinter.CTkLabel(self.janela, text=(
                "ERRO\nUsuário não cadastrado!!!"), text_color="red", font=("Century Gothic Bold", 11))
            erro_msg.pack(padx=10, pady=5)
            erro_msg.after(4000, erro_msg.destroy)







    def cadastrar(self):
        self.usuario_cadastrado = self.Usuario.get()
        self.senha_digitada = self.senha.get()
        # Verifique se o usuário já existe no banco de dados
        self.cursor.execute("SELECT * FROM usuarios WHERE usuario=?",
                            (self.usuario_cadastrado,))
        self.resultado = self.cursor.fetchone()

        if self.usuario_cadastrado.strip() == "" or self.senha_digitada.strip() == "":
            mensagem_cadastro = customtkinter.CTkLabel(self.janela, text=(
                "Preencha todos os espaços vazios!!"), font=("Century Gothic Bold", 10), text_color=("Red"))
            mensagem_cadastro.pack(padx=10, pady=5)
            mensagem_cadastro.after(2000, mensagem_cadastro.destroy)

        else:
            # Insira o novo usuário no banco de dados
            self.cursor.execute("INSERT INTO usuarios (usuario, senha) VALUES (?, ?)",
                                (self.usuario_cadastrado.lower(), self.senha_digitada.lower()))
            self.conn.commit()
            self.mensagem_cadastro2 = customtkinter.CTkLabel(self.janela, text=(
                f"Conta Cadastrada!!\nBem Vindo ao nosso time\n{self.usuario_cadastrado.upper()}"), font=("Century Gothic Bold", 15), text_color=("green"))
            self.mensagem_cadastro2.pack(padx=10, pady=5)
            self.mensagem_cadastro2.after(
                4000, self.mensagem_cadastro2.destroy)
            self.botao_cadastro.destroy()


    def menu(self):
        self.canvas_funcionarios_criado = False
        self.canvas_criado = False
        self.menuP = customtkinter.CTk()
        largura = 1360
        altura = 720
        
        self.menuP.title("ObraPlus - Menu Principal")
        
        self.menuP.iconbitmap('Obra_1_.ico')
        
        self.menuP.geometry(f"{largura}x{altura}")
        self.menuP._set_appearance_mode("dark")

        # Obtém as dimensões da tela
        largura_tela = self.menuP.winfo_screenwidth()
        altura_tela = self.menuP.winfo_screenheight()

        # Calcula a posição central da janela
        x = (largura_tela - largura) // 2
        y = (altura_tela - altura) // 2

        # Define a geometria da janela para centralizá-la
        self.menuP.geometry(f"{largura}x{altura}+{x}+{y}")



        


        #criando menu arte
        self.welcome_label = customtkinter.CTkLabel(self.menuP, text=" Bem-vindo ao ObraPlus", font=("Sedgwick Ave Display ", 30), text_color="#7B68EE")  
        self.welcome_label.place(relx=0.45, rely=0.02)  


        # Cria uma Label dentro do Canvas
        self.label = customtkinter.CTkLabel(self.menuP, text="ObraPlus",font=("Sedgwick Ave Display ", 25),text_color="#9370DB")
        self.label.place(relx=0.58, rely=0.11, anchor="center")
        
        
        self.label = customtkinter.CTkLabel(self.menuP, text=" É uma ferramenta definitiva para simplificar e otimizar\n\no gerenciamento de projetos de construção\n\n permitindo que você tenha o controle total de suas obras\n\n de forma eficaz e conveniente.",font=("Sedgwick Ave Display ", 15))
        self.label.place(relx=0.58, rely=0.22, anchor="center")


        self.label2 = customtkinter.CTkLabel(self.menuP, text="Funcionalidade",font=("Sedgwick Ave Display ", 25),text_color="#9370DB")
        self.label2.place(relx=0.58, rely=0.34, anchor="center")
        
        self.label2 = customtkinter.CTkLabel(self.menuP, text="Cadastro de Clientes,Funcionários e Etapas da Obra:\n\nSimplifique o processo de gerenciamento de pessoas com o cadastro de clientes e funcionários.\n\nRegistre informações detalhadas, como nomes, contatos, funções e informações \n\nrelevantes sobre cada indivíduo\n\nAcompanhe o progresso de suas obras de construção de forma sistemática com a funcionalidade de etapas da obra.\n\n Divida seu projeto em fases específicas\n\n como preparação do terreno, fundação, estrutura, acabamento, entre outras",font=("Sedgwick Ave Display ", 15))
        self.label2.place(relx=0.58, rely=0.54, anchor="center")

        self.label3 = customtkinter.CTkLabel(self.menuP, text="Suporte e Ajuda",font=("Sedgwick Ave Display ", 25),text_color="#9370DB")
        self.label3.place(relx=0.58, rely=0.73, anchor="center")
        
        self.label3 = customtkinter.CTkLabel(self.menuP, text="Precisa de ajuda ou tem alguma dúvida?\n\nNosso suporte está à disposição para ajudá-lo a aproveitar ao máximo o Obra Plus. \n\nEntre em contato através do whatsapp :\n\n(31)98703-5797 Lucas guerra\n\n(31)97544-9265 Gustavo Henrique ",font=("Sedgwick Ave Display ", 15))
        self.label3.place(relx=0.58, rely=0.87, anchor="center")
       
       
       
       
        # criar Frame(quadrado cinza):
        self.frame = customtkinter.CTkFrame(
            master=self.menuP, width=270, height=1080)
        self.frame.pack(side=LEFT)

        # NomeAPP:
        self.fontobra = customtkinter.CTkLabel(self.frame, text="ObraPlus", font=(
            "Press Start 2P", 45), text_color="#9400D3")
        self.fontobra.place(x=50, y=10)

        # NomeMenu:
        self.fontobra = customtkinter.CTkLabel(self.frame, text="Menu Principal", font=(
            "Press Start 2P", 20), text_color="#f3f3ea")
        self.fontobra.place(x=70, y=70)



        # botao cliente:
        self.cliente = customtkinter.CTkButton(
            self.frame, text="Clientes", font=("Press Start 2P", 30), text_color="#fff", fg_color="transparent", hover_color="#9400D3",width=270, height=30, command=self.botao_clientes).place(y=200)

        # botao Funcionario:
        self.funcionario = customtkinter.CTkButton(
            self.frame, text="Funcionarios", font=("Press Start 2P", 30), text_color="#fff", fg_color="transparent", hover_color="#9400D3", width=270, height=30, command=self.botao_funcionarios).place(y=290)

        # botao etapas:
        self.etapas = customtkinter.CTkButton(
            self.frame, text="Etapas", font=("Press Start 2P", 30), text_color="#fff", fg_color="transparent", hover_color="#9400D3", width=270, height=30,command=self.botao_etapas).place(y=390)
        
        self.etapas = customtkinter.CTkButton(
            self.frame, text="Etapas", font=("Press Start 2P", 30), text_color="#fff", fg_color="transparent", hover_color="#9400D3", width=270, height=30,command=self.botao_etapas).place(y=390)

        # NomeContato:
        self.fontobra = customtkinter.CTkLabel(self.frame, text="Contato📞:" ,font=(
            "Press Start 2P", 25), text_color="#f3f3ea")
        self.fontobra.place(x=5, rely=0.90,anchor ='sw')


        self.fontobra = customtkinter.CTkLabel(self.frame, text="(31)98703-5797 Lucas Guerra",font=(
            "Press Start 2P", 15), text_color="#f3f3ea")
        self.fontobra.place(x=5, rely=0.95,anchor ='sw')


        
        self.fontobra = customtkinter.CTkLabel(self.frame, text="(31)97544-9265 Gustavo Henrique",font=(
            "Press Start 2P", 15), text_color="#f3f3ea")
        self.fontobra.place(x=5, rely=1.0,anchor ='sw')







        # botao sair
        #self.sair = customtkinter.CTkButton(
            #self.frame, text="Sair", font=("Press Start 2P", 30), text_color="#fff", fg_color="#8b0000", hover_color="#640b0b", height=30, command=self.log_off)
        #self.sair.place(x=50, rely=1.0, anchor='sw')
        
        self.menuP.mainloop()

    def abrir_menu(self):
        self.conn.close()
        self.janela.withdraw()
        self.menu()
   
    #def log_off(self):
        self.menuP.withdraw()  # Esconda a janela do menu
        self.janela.destroy()   # Destrua a janela de login atual
        sistema = TelaLogin()   
    
   
   
   
   
    # Função para cadastrar um cliente
    def cadastrar_cliente(self):
        nome = self.nome_entry.get()
        idade = self.idade_entry.get()
        endereco = self.endereco_entry.get()
        bairro = self.bairro_entry.get()
        numero_endereco = self.numero_entry.get()
        telefone = self.telefone_entry.get()
        wpp = self.wpp_entry.get()
        cpf = self.cpf_entry.get()
        email = self.email_entry.get()
        OBs = self.OBs_entry.get()


        if len(cpf) != 11 or not cpf.isdigit():
            messagebox.showerror("Erro", "CPF inválido. Deve conter 11 dígitos numéricos.")
            return

        if not re.match(r'^\d{11}$', telefone):
            messagebox.showerror("Erro", "Telefone inválido. Deve conter 11 dígitos numéricos.")
            return

        if not re.match(r'^\d{11}$', wpp):
            messagebox.showerror("Erro", "Telefone inválido. Deve conter 11 dígitos numéricos.")
            return

        if not idade.isdigit():
            messagebox.showerror("Erro", "Idade inválida. Deve conter apenas números.")
            return

        try:
            # Verifica se a pasta "clientes" existe e cria-a se não existir
            if not os.path.exists("clientes"):
                os.makedirs("clientes")

            # Caminho completo para a planilha de clientes dentro da pasta "clientes"
            workbook_path = os.path.join("clientes", 'clientes.xlsx')

            try:
                workbook = openpyxl.load_workbook(workbook_path)
                sheet = workbook.active
            except FileNotFoundError:
                workbook = openpyxl.Workbook()
                sheet = workbook.active
                sheet.append(["Nome", "Idade", "Endereço", "Numero endereço", "Bairro", "Telefone", "WPP", "CPF", "E-mail", "Observação"])

            for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, values_only=True):
                if row[7] == cpf:
                    messagebox.showerror("Erro", "CPF já cadastrado.")
                    return

            sheet.append([nome, idade, endereco, numero_endereco, bairro, telefone, wpp, cpf, email, OBs])
            workbook.save(workbook_path)
            messagebox.showinfo("Sucesso", "Cliente cadastrado com sucesso!")
        except Exception as e:
            messagebox.showerror("Erro", f"Ocorreu um erro: {str(e)}")


        # Limpa os campos de entrada após o cadastro bem-sucedido
        self.nome_entry.delete(0, 'end')
        self.idade_entry.delete(0, 'end')
        self.endereco_entry.delete(0, 'end')
        self.numero_entry.delete(0, 'end')
        self.telefone_entry.delete(0, 'end')
        self.cpf_entry.delete(0, 'end')
        self.wpp_entry.delete(0, 'end')
        self.bairro_entry.delete(0, 'end')
        self.email_entry.delete(0, 'end')
        self.OBs_entry.delete(0,'end')



    # Função para abrir a janela de clientes
    def botao_clientes(self):
        if not self.botao_clicado:
            self.botao_clicado = True
        if self.canvas_funcionarios and self.canvas_funcionarios.winfo_ismapped():
            self.canvas_funcionarios.pack_forget()
        if self.canvas_etapas and self.canvas_etapas.winfo_ismapped():
            self.canvas_etapas.pack_forget()
        if self.tela_atual != "clientes":
            self.mostrar_tela_clientes()
            self.tela_atual = "clientes"

            

    
    def mostrar_tela_clientes(self):
                
                
                self.canvas_cliente = customtkinter.CTkCanvas(self.menuP, bg="#303030")
                self.canvas_cliente.pack(fill=customtkinter.BOTH, expand=True)

                self.nome_label = customtkinter.CTkLabel(self.canvas_cliente, text="Nome:", font=("Press Start 2P", 30), text_color="#fff")
                
                self.idade_label = customtkinter.CTkLabel(self.canvas_cliente, text="Idade:", font=("Press Start 2P", 30), text_color="#fff")
                
                self.endereco_label = customtkinter.CTkLabel(self.canvas_cliente, text="Endereço:", font=("Press Start 2P", 30), text_color="#fff")
                                
                self.bairro_label = customtkinter.CTkLabel(self.canvas_cliente, text="Bairro: ", font=("Press Start 2P", 30), text_color="#fff")
                                
                self.numero_label = customtkinter.CTkLabel(self.canvas_cliente, text="Nº:", font=("Press Start 2P", 30), text_color="#fff")

                self.email_label = customtkinter.CTkLabel(self.canvas_cliente, text="E-mail:", font=("Press Start 2P", 30), text_color="#fff")
                
                self.telefone_label = customtkinter.CTkLabel(self.canvas_cliente, text="Telefone:", font=("Press Start 2P", 30), text_color="#fff")

                self.wpp_label = customtkinter.CTkLabel(self.canvas_cliente, text="Numero do Wpp: ", font=("Press Start 2P", 30), text_color="#fff")
                
                self.cpf_label = customtkinter.CTkLabel(self.canvas_cliente, text="CPF:", font=("Press Start 2P", 30), text_color="#fff")

                self.OBs_label = customtkinter.CTkLabel(self.canvas_cliente, text="Observação:", font=("Press Start 2P", 30), text_color="#fff")







                self.recado_label = customtkinter.CTkLabel(self.canvas_cliente, text="Bem-vindo ao nosso cadastro de cliente", font=("Helvetica", 40), text_color="#7B68EE")
                
                #subTitle
                self.subti_label = customtkinter.CTkLabel(self.canvas_cliente, text="➦Preencha todos os campos vazios !", font=("Helvetica", 20), text_color="#fff")





                self.nome_entry = customtkinter.CTkEntry(self.canvas_cliente,width=250,height=35,font=('Helvetica', 20),fg_color="transparent")
                
          
                self.idade_entry = customtkinter.CTkEntry(self.canvas_cliente,width=250,height=35,font=('Helvetica', 20),fg_color="transparent")


                self.bairro_entry = customtkinter.CTkEntry(self.canvas_cliente,width=250,height=35,font=('Helvetica', 20),fg_color="transparent")
         
                self.endereco_entry = customtkinter.CTkEntry(self.canvas_cliente,width=250,height=35,font=('Helvetica', 20),fg_color="transparent")

                self.numero_entry = customtkinter.CTkEntry(self.canvas_cliente,width=100,height=35,font=('Helvetica', 30),fg_color="transparent")
                                
                self.email_entry = customtkinter.CTkEntry(self.canvas_cliente,width=250,height=35,font=('Helvetica', 20),fg_color="transparent")

                self.telefone_entry = customtkinter.CTkEntry(self.canvas_cliente,width=250,height=35,font=('Helvetica', 20),fg_color="transparent")

                self.wpp_entry = customtkinter.CTkEntry(self.canvas_cliente,width=250,height=35,font=('Helvetica', 20) ,placeholder_text="com digito 9",fg_color="transparent")
         
                self.cpf_entry = customtkinter.CTkEntry(self.canvas_cliente,width=250,height=35,font=('Helvetica', 20) , placeholder_text="apenas números !",fg_color="transparent")
                
                self.OBs_entry = customtkinter.CTkEntry(self.canvas_cliente,width=250,height=35,font=('Helvetica', 20) ,fg_color="transparent")
                

                

                
                                
                #botao cadastro
                self.cadastrar_button = customtkinter.CTkButton(self.canvas_cliente, text="Cadastrar meu Cliente", font=("Press Start 2P", 30), text_color="#000",fg_color="#9370DB", hover_color="#4B0082" ,command=self.cadastrar_cliente, width=30, height=30)












                self.nome_label.place(relx=0.25, rely=0.20)
                self.nome_entry.place(relx=0.25, rely=0.25)

                self.idade_label.place(relx=0.25, rely=0.30)
                self.idade_entry.place(relx=0.25, rely=0.35)


                self.bairro_label.place(relx=0.25, rely=0.40)
                self.bairro_entry.place(relx=0.25, rely=0.45)

                self.endereco_label.place(relx=0.25, rely=0.50)
                self.endereco_entry.place(relx=0.25, rely=0.55)
                
                self.numero_label.place(relx=0.25, rely=0.60)
                self.numero_entry.place(relx=0.25, rely=0.65)

                
                self.email_label.place(relx=0.50, rely=0.20)
                self.email_entry.place(relx=0.50, rely=0.25)

                self.telefone_label.place(relx=0.50, rely=0.30)
                self.telefone_entry.place(relx=0.50, rely=0.35)


                self.wpp_label.place(relx=0.50, rely=0.40)
                self.wpp_entry.place(relx=0.50, rely=0.45)

                self.cpf_label.place(relx=0.50, rely=0.50)
                self.cpf_entry.place(relx=0.50, rely=0.55)



                self.OBs_label.place(relx=0.50, rely=0.60)
                self.OBs_entry.place(relx=0.50, rely=0.65)





                self.recado_label.place(relx=0.20,rely=0.02)

                self.subti_label.place(relx=0.20,rely=0.09)

                
                self.cadastrar_button.place(relx=0.34, rely=0.94)











    # Função para cadastrar um funcionário
    def cadastrar_funcionario(self):
        nome = self.nome_entry.get()
        idade = self.idade_entry.get()
        endereco = self.endereco_entry.get()
        bairro = self.bairro_entry.get()
        numero_endereco = self.numero_entry.get()
        telefone = self.telefone_entry.get()
        wpp = self.wpp_entry.get()
        cpf = self.cpf_entry.get()
        email = self.email_entry.get()
        profissao = self.profissao_entry.get()
        funcao = self.funcao_combobox.get()
        obs = self.obs_entry.get()

        if len(cpf) != 11 or not cpf.isdigit():
            messagebox.showerror("Erro", "CPF inválido. Deve conter 11 dígitos numéricos.")
            return

        if not re.match(r'^\d{11}$', telefone):
            messagebox.showerror("Erro", "Telefone inválido. Deve conter 11 dígitos numéricos.")
            return

        if not re.match(r'^\d{11}$', wpp):
            messagebox.showerror("Erro", "Telefone inválido. Deve conter 11 dígitos numéricos.")
            return

        if not idade.isdigit():
            messagebox.showerror("Erro", "Idade inválida. Deve conter apenas números.")
            return

        try:
            # Verifica se a pasta "funcionarios" existe e cria-a se não existir
            if not os.path.exists("funcionarios"):
                os.makedirs("funcionarios")

            # Caminho completo para a planilha de funcionários dentro da pasta "funcionarios"
            workbook_path = os.path.join("funcionarios", 'funcionarios.xlsx')

            try:
                workbook = openpyxl.load_workbook(workbook_path)
                sheet = workbook.active
            except FileNotFoundError:
                workbook = openpyxl.Workbook()
                sheet = workbook.active
                sheet.append(["Nome", "Idade", "Endereço", "Numero endereço", "Bairro", "Telefone", "WPP", "CPF", "E-mail", "Profissão", "Função", "Observação"])

            for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, values_only=True):
                if row[7] == cpf:
                    messagebox.showerror("Erro", "CPF já cadastrado.")
                    return

            sheet.append([nome, idade, endereco, numero_endereco, bairro, telefone, wpp, cpf, email, profissao, funcao,obs])
            workbook.save(workbook_path)
            messagebox.showinfo("Sucesso", "Funcionário cadastrado com sucesso!")
        except Exception as e:
            messagebox.showerror("Erro", f"Ocorreu um erro: {str(e)}")




        # Limpa os campos de entrada após o cadastro bem-sucedido
        self.nome_entry.delete(0, 'end')
        self.idade_entry.delete(0, 'end')
        self.endereco_entry.delete(0, 'end')
        self.numero_entry.delete(0, 'end')
        self.telefone_entry.delete(0, 'end')
        self.cpf_entry.delete(0, 'end')
        self.wpp_entry.delete(0, 'end')
        self.bairro_entry.delete(0, 'end')
        self.email_entry.delete(0, 'end')
        self.profissao_entry.delete(0, 'end')
        self.obs_entry.delete(0,'end')
        

    # Função para abrir a janela de funcionários
    def botao_funcionarios(self):
        if not self.botao_clicado:
            self.botao_clicado = True
        
        if self.canvas_cliente and self.canvas_cliente.winfo_ismapped():
            self.canvas_cliente.pack_forget()
        
        if self.canvas_etapas and self.canvas_etapas.winfo_ismapped():
            self.canvas_etapas.pack_forget()

        if self.tela_atual != "funcionarios":
            self.mostrar_tela_funcionarios()
            self.tela_atual = "funcionarios"

                


                                    

    def mostrar_tela_funcionarios(self):
        
                self.canvas_funcionarios = customtkinter.CTkCanvas(self.menuP, bg="#303030")
                self.canvas_funcionarios.pack(fill=customtkinter.BOTH, expand=True)



                self.nome_label = customtkinter.CTkLabel(self.canvas_funcionarios, text="Nome:", font=("Press Start 2P", 30), text_color="#fff")
                self.idade_label = customtkinter.CTkLabel(self.canvas_funcionarios, text="Idade:", font=("Press Start 2P", 30), text_color="#fff")
                self.endereco_label = customtkinter.CTkLabel(self.canvas_funcionarios, text="Endereço:", font=("Press Start 2P", 30), text_color="#fff")
                self.numero_label = customtkinter.CTkLabel(self.canvas_funcionarios, text="Nº:", font=("Press Start 2P", 30), text_color="#fff")
                self.telefone_label = customtkinter.CTkLabel(self.canvas_funcionarios, text="Telefone:", font=("Press Start 2P", 30), text_color="#fff")
                self.wpp_label = customtkinter.CTkLabel(self.canvas_funcionarios, text="Numero do Wpp: ", font=("Press Start 2P", 30), text_color="#fff")
                self.cpf_label = customtkinter.CTkLabel(self.canvas_funcionarios, text="CPF:", font=("Press Start 2P", 30), text_color="#fff")
                self.bairro_label = customtkinter.CTkLabel(self.canvas_funcionarios, text="Bairro: ", font=("Press Start 2P", 30), text_color="#fff")
                self.email_label = customtkinter.CTkLabel(self.canvas_funcionarios, text="E-mail:", font=("Press Start 2P", 30), text_color="#fff")
                self.profissao_label = customtkinter.CTkLabel(self.canvas_funcionarios, text="Profissão:", font=("Press Start 2P", 30), text_color="#fff")
                self.funcao_label = customtkinter.CTkLabel(self.canvas_funcionarios, text="Função:", font=("Press Start 2P", 30), text_color="#fff")
                self.obs_label = customtkinter.CTkLabel(self.canvas_funcionarios, text="Observação:", font=("Press Start 2P", 30), text_color="#fff")

                
                
                #TITTLE

                self.recado1_label = customtkinter.CTkLabel(self.canvas_funcionarios, text="Bem-vindo ao nosso cadastro de funcionarios", font=("Helvetica", 40), text_color="#7B68EE")
                
                #subTitle
                self.subti1_label = customtkinter.CTkLabel(self.canvas_funcionarios, text="➦Preencha todos os campos vazios !", font=("Helvetica", 20), text_color="#fff")
                
                
                self.nome_entry = customtkinter.CTkEntry(self.canvas_funcionarios,width=250,height=35,font=('Helvetica', 20),fg_color="transparent")
                self.idade_entry = customtkinter.CTkEntry(self.canvas_funcionarios,width=250,height=35,font=('Helvetica', 20),fg_color="transparent")
                self.bairro_entry = customtkinter.CTkEntry(self.canvas_funcionarios,width=250,height=35,font=('Helvetica', 20),fg_color="transparent")
                self.endereco_entry = customtkinter.CTkEntry(self.canvas_funcionarios,width=250,height=35,font=('Helvetica', 20),fg_color="transparent")
                self.numero_entry = customtkinter.CTkEntry(self.canvas_funcionarios,width=100,height=35,font=('Helvetica', 20),fg_color="transparent")
                
                self.funcao_combobox = customtkinter.CTkComboBox(self.canvas_funcionarios, values=["Supervisor", "Planejamento", "Comprador", "Orçamentista", "Qualidade"], font=('Helvetica', 20))
                self.funcao_combobox.set("Supervisor")
                
                self.telefone_entry = customtkinter.CTkEntry(self.canvas_funcionarios,width=250,height=35,font=('Helvetica', 20),fg_color="transparent")
                self.wpp_entry = customtkinter.CTkEntry(self.canvas_funcionarios,width=250,height=35,font=('Helvetica', 20),fg_color="transparent")
                self.email_entry = customtkinter.CTkEntry(self.canvas_funcionarios,width=250,height=35,font=('Helvetica', 20),fg_color="transparent")
                self.profissao_entry = customtkinter.CTkEntry(self.canvas_funcionarios,width=250,height=35,font=('Helvetica', 20),fg_color="transparent")
                self.cpf_entry = customtkinter.CTkEntry(self.canvas_funcionarios,width=250,height=35,font=('Helvetica', 20),fg_color="transparent")
                self.obs_entry = customtkinter.CTkEntry(self.canvas_funcionarios,width=250,height=35,font=('Helvetica', 20),fg_color="transparent")
                
                self.funcao_combobox = customtkinter.CTkComboBox(self.canvas_funcionarios, values=["Supervisor", "Planejamento", "Comprador", "Orçamentista", "Qualidade"])
                self.funcao_combobox.set("Supervisor")

                self.cadastrar_button = customtkinter.CTkButton(self.canvas_funcionarios, text="Cadastrar meu funcionario",  font=("Press Start 2P", 30), text_color="#000",fg_color="#9370DB", hover_color="#4B0082" ,command=self.cadastrar_funcionario, width=30, height=30)








                self.nome_label.place(relx=0.25, rely=0.20)
                self.nome_entry.place(relx=0.25, rely=0.25)

                self.idade_label.place(relx=0.25, rely=0.30)
                self.idade_entry.place(relx=0.25, rely=0.35)

                self.bairro_label.place(relx=0.25, rely=0.40)
                self.bairro_entry.place(relx=0.25, rely=0.45)

                self.endereco_label.place(relx=0.25, rely=0.50)
                self.endereco_entry.place(relx=0.25, rely=0.55)



                self.numero_label.place(relx=0.25, rely=0.60)
                self.numero_entry.place(relx=0.25, rely=0.65)

                self.funcao_label.place(relx=0.25, rely=0.70)
                self.funcao_combobox.place(relx=0.25, rely=0.75)

                self.telefone_label.place(relx=0.50, rely=0.20)
                self.telefone_entry.place(relx=0.50, rely=0.25)


                self.wpp_label.place(relx=0.50, rely=0.30)
                self.wpp_entry.place(relx=0.50, rely=0.35)

                self.email_label.place(relx=0.50, rely=0.40)
                self.email_entry.place(relx=0.50, rely=0.45)


                self.profissao_label.place(relx=0.50, rely=0.50)
                self.profissao_entry.place(relx=0.50, rely=0.55)


                self.cpf_label.place(relx=0.50, rely=0.60)
                self.cpf_entry.place(relx=0.50, rely=0.65)


                self.obs_label.place(relx=0.50, rely=0.70)
                self.obs_entry.place(relx=0.50, rely=0.75)

                



                
                self.recado1_label.place(relx=0.20,rely=0.02)

                self.subti1_label.place(relx=0.20,rely=0.09)



                self.cadastrar_button.place(relx=0.32, rely=0.94)






    # Função para abrir a janela de funcionários
    def botao_etapas(self):
        if not self.botao_clicado:
            self.botao_clicado = True
        if self.canvas_cliente and self.canvas_cliente.winfo_ismapped():
            self.canvas_cliente.pack_forget()
        if self.canvas_funcionarios and self.canvas_funcionarios.winfo_ismapped():
            self.canvas_funcionarios.pack_forget()
        if self.tela_atual != "etapas":
            self.mostrar_tela_etapas()
            self.tela_atual = "etapas"



            # Validation functions
    def validate_duração(self, P):
        if P == "" or P.isdigit():
            return True
        else:
            return False

    def validate_valor_orcado(self, P):
        if P == "" or (P.replace(".", "", 1).isdigit() and P.count(".") <= 1):
            return True
        else:
            return False

    def validate_por_m(self, P):
        if P == "" or P.isdigit():
            return True
        else:
            return False
            

    def mostrar_tela_etapas(self):

    
                
        
                self.canvas_etapas = customtkinter.CTkCanvas(self.menuP, bg="#303030")
                self.canvas_etapas.pack(fill=customtkinter.BOTH, expand=True)


                self.createEtp_label = customtkinter.CTkLabel(self.canvas_etapas, text="Crie uma nova etapa", font=("Press Start 2P",40), text_color="#7B68EE")
                self.createEtp_label.place(relx=0.35, rely=0.02)
                
                self.subti3_label = customtkinter.CTkLabel(self.canvas_etapas, text="➦Preencha todos os campos vazios !", font=("Helvetica", 20), text_color="#fff")
                self.subti3_label.place(relx=0.35,rely=0.09)



                #LABELS E ENTRYS

                self.newETP_label = customtkinter.CTkLabel(self.canvas_etapas,text="Nome da Etapa :",font=('Helvetica', 30),fg_color="transparent")
                self.newETP_entry = customtkinter.CTkEntry(self.canvas_etapas,width=250,height=35,font=('Helvetica', 30),fg_color="transparent")
                self.newETP_label.place(relx=0.25, rely=0.20)
                self.newETP_entry.place(relx=0.25, rely=0.25)

                self.responseName_label = customtkinter.CTkLabel(self.canvas_etapas,text="Responsável :",font=('Helvetica', 30))
                self.responseName_entry = customtkinter.CTkEntry(self.canvas_etapas,width=250,height=35,font=('Helvetica', 30),fg_color="transparent")
                self.responseName_label.place(relx=0.25, rely=0.35)
                self.responseName_entry.place(relx=0.25, rely=0.40)


                self.inicio_label = customtkinter.CTkLabel(self.canvas_etapas, text="Data de Inicio:", font=("Press Start 2P", 30), text_color="#fff")
                self.inicio_label.place(relx=0.25, rely=0.50)
                self.inicio_calendar_entry = customtkinter.CTkEntry(self.canvas_etapas,width=250,height=35, font=('Helvetica', 30), placeholder_text="\ndd/mm/yyyy", fg_color="transparent")
                self.inicio_calendar_entry.place(relx=0.25, rely=0.55)
                

                self.duracao_label = customtkinter.CTkLabel(self.canvas_etapas, text="Duração (dias):", font=("Press Start 2P", 30), text_color="#fff")
                self.duracao_label.place(relx=0.25, rely=0.65)
                self.duração_calendar_entry = customtkinter.CTkEntry(self.canvas_etapas, width=100, height=35, font=('Helvetica', 30), fg_color="transparent")
                self.duração_calendar_entry.place(relx=0.25, rely=0.70)
                self.duração_calendar_entry.configure(validate="key", validatecommand=(self.duração_calendar_entry.register(self.validate_duração), "%P"))
                self.result_label = customtkinter.CTkLabel(self.canvas_etapas, text="", font=("Press Start 2P", 20), text_color="#fff")
                self.result_label.place(relx=0.20, rely=0.77)


                self.nomeobra_label = customtkinter.CTkLabel(self.canvas_etapas, text="Nomeia a Obra :", font=("Press Start 2P",30), text_color="#fff")
                self.nomeobra_label.place(relx=0.50, rely=0.20)
                self.nomeobra_entry = customtkinter.CTkEntry(self.canvas_etapas,width=250,height=35,font=('Helvetica', 30),fg_color="transparent")
                self.nomeobra_entry.place(relx=0.50 ,rely=0.25)


                self.valor_orcado_label = customtkinter.CTkLabel(self.canvas_etapas, text="Valor Orçado (R$):", font=("Press Start 2P", 30), text_color="#fff")
                self.valor_orcado_label.place(relx=0.50, rely=0.35)
                self.valor_orcado_entry = customtkinter.CTkEntry(self.canvas_etapas, width=250, height=35, font=('Helvetica', 30), fg_color="transparent")
                self.valor_orcado_entry.place(relx=0.50, rely=0.40)
                self.valor_orcado_entry.configure(validate="key", validatecommand=(self.valor_orcado_entry.register(self.validate_valor_orcado), "%P"))



                self.por_M_label = customtkinter.CTkLabel(self.canvas_etapas, text="Qtd/Medida :", font=("Press Start 2P",30), text_color="#fff")
                self.por_M_label.place(relx=0.50, rely=0.50)
                self.por_M_entry = customtkinter.CTkEntry(self.canvas_etapas, width=100, height=35, font=('Helvetica', 30), fg_color="transparent")
                self.por_M_entry.place(relx=0.50, rely=0.55)
                self.por_M_entry.configure(validate="key", validatecommand=(self.por_M_entry.register(self.validate_por_m), "%P"))



                self.por_label = customtkinter.CTkLabel(self.canvas_etapas, text="% Etapa :", font=("Press Start 2P", 30), text_color="#fff")
                self.por_label.place(relx=0.50, rely=0.65)
                self.por_entry = customtkinter.CTkEntry(self.canvas_etapas,width=250,height=35,font=('Helvetica', 30),fg_color="transparent")
                self.por_entry.place(relx=0.50, rely=0.70)


                opcoes = [" Metro quadrado (m²)", " Metro cúbico (m³)", " Item (it.)", " Unidade (un.)","Verba (vb.)"]
                self.comboBox = customtkinter.CTkComboBox(self.canvas_etapas, values=opcoes, font=('Helvetica', 20),width=230)
                self.comboBox.place(relx=0.50, rely=0.80)
                self.comboBox.set("Unidade de cálculo")



                #botoes


                self.calcular_button = customtkinter.CTkButton(self.canvas_etapas, fg_color="#9370DB", hover_color="#4B0082" ,text="Calcular Prazo", font=("Press Start 2P", 20),text_color="#000", command=self.calcular_data_conclusao)
                self.calcular_button.place(relx=0.35, rely=0.93)    

                self.salvar_button = customtkinter.CTkButton(self.canvas_etapas, fg_color="#9370DB", hover_color="#4B0082" ,text="Salvar", font=("Press Start 2P", 20),text_color="#000", command=self.salvar_etapa)
                self.salvar_button.place(relx=0.50, rely=0.93)









    def calcular_porcentagem_etapa(self):
        por = self.por_entry.get()
        valor_orcado = self.valor_orcado_entry.get()

        # Verifique se por e valor_orcado não estão vazios antes de tentar a conversão
        if por and valor_orcado:
            try:
                valor_etapa = float(por)
                valor_total = float(valor_orcado)
                porcentagem_etapa = (valor_etapa / valor_total) * 100
                return porcentagem_etapa
            except ValueError:
                messagebox.showerror("Erro", "Valores inválidos para por ou valor_orcado.")
        else:
            messagebox.showerror("Erro", "Por e/ou valor_orcado não podem estar vazios.")
        return None


    def salvar_etapa(self):
        nome_etapa = self.newETP_entry.get()
        responsavel = self.responseName_entry.get()
        qtd_medida = self.por_M_entry.get()
        unidade_calculo = self.comboBox.get()
        data_inicio = self.inicio_calendar_entry.get()
        duracao = self.duração_calendar_entry.get()
        nome_da_obra = self.nomeobra_entry.get()
        valor_orcado = self.valor_orcado_entry.get()
        porcentagem_etapa = self.por_entry.get()

        if not porcentagem_etapa:
            messagebox.showerror("Erro", "Porcentagem da etapa não pode estar vazia.")
            return

        # Verifica se já existe uma etapa com o mesmo nome da obra
        for etapa in self.etapas_data:
            if etapa["Nome da Obra"] == nome_da_obra:
                # Atualiza os dados da etapa existente em vez de adicionar uma nova entrada
                etapa.update({
                    "Nome da Etapa": nome_etapa,
                    "Responsável": responsavel,
                    "Qtd/Medida": qtd_medida,
                    "Unidade de Cálculo": unidade_calculo,
                    "Data de Início": data_inicio,
                    "Duração (dias)": duracao,
                    "Nome da Obra": nome_da_obra,
                    "Valor Orçado": valor_orcado,
                    "% ETP": float(porcentagem_etapa),
                })
                break
        else:
            # Se não encontrou uma etapa com o mesmo nome da obra, adiciona uma nova entrada
            self.etapas_data.append({
                "Nome da Etapa": nome_etapa,
                "Responsável": responsavel,
                "Qtd/Medida": qtd_medida,
                "Unidade de Cálculo": unidade_calculo,
                "Data de Início": data_inicio,
                "Duração (dias)": duracao,
                "Nome da Obra": nome_da_obra,
                "Valor Orçado": valor_orcado,
                "% ETP": float(porcentagem_etapa),
            })

        # Limpe os campos de entrada, exceto o campo "Nome da Obra"
        self.newETP_entry.delete(0, tk.END)
        self.responseName_entry.delete(0, tk.END)
        self.por_M_entry.delete(0, tk.END)
        self.inicio_calendar_entry.delete(0, tk.END)
        self.duração_calendar_entry.delete(0, tk.END)
        self.valor_orcado_entry.delete(0, tk.END)
        self.por_entry.delete(0, tk.END)

        # Atualize a tabela Excel
        self.atualizar_tabela_excel()

        # Atualize a mensagem na label result_label
        self.result_label.configure(text="Dados salvos com sucesso!", text_color="#00ff00")
        self.result_label.after(1500, self.limpar_mensagem)







    def limpar_mensagem(self):
        # Limpe o texto da label result_label
        self.result_label.configure(text="")

   
   
   
    def atualizar_tabela_excel(self):
        # Atualiza a tabela Excel
        for etapa in self.etapas_data:
            porcentagem_etapa = etapa["% ETP"]
            for outra_etapa in self.etapas_data:
                if etapa["Nome da Obra"] == outra_etapa["Nome da Obra"]:
                    porcentagem_etapa += outra_etapa["% ETP"]
            etapa["% Obra Total"] = porcentagem_etapa

        # Cria um DataFrame a partir da lista etapas_data
        df = pd.DataFrame(self.etapas_data)

        # Agrupa os dados por "Nome da Obra"
        grupos = df.groupby("Nome da Obra")

        for nome_obra, grupo in grupos:
            # Cria a pasta com o nome da obra
            pasta_obra = os.path.join(os.getcwd(), nome_obra)
            if not os.path.exists(pasta_obra):
                os.mkdir(pasta_obra)

            # Cria o caminho completo para a planilha dentro da pasta
            caminho_planilha = os.path.join(pasta_obra, f"{nome_obra}_Etapas.xlsx")

            # Verifica se a planilha já existe
            if os.path.exists(caminho_planilha):
                # Carrega a planilha existente em um DataFrame do Pandas
                df_existente = pd.read_excel(caminho_planilha)

                # Concatena o DataFrame existente com as novas etapas
                df_existente = pd.concat([df_existente, grupo], ignore_index=True)
            else:
                # Se a planilha não existir, cria um DataFrame com as etapas
                df_existente = grupo

            # Salva o DataFrame de volta na mesma planilha existente
            df_existente.to_excel(caminho_planilha, sheet_name='Etapas', index=False)

            # Atualiza o relatório em PDF
            self.atualizar_relatorio_pdf(nome_obra, df_existente)

        # Cria a mensagem após o loop
        mensagem = "Tabelas Excel atualizadas com sucesso para as obras!"
        messagebox.showinfo("Sucesso", mensagem)




    def atualizar_relatorio_pdf(self, nome_obra, df):
        # Cria o caminho completo para o arquivo PDF
        caminho_pdf = os.path.join(os.getcwd(), nome_obra, f"{nome_obra}_Relatorio.pdf")

        # Cria um documento PDF usando a biblioteca reportlab
        c = canvas.Canvas(caminho_pdf, pagesize=letter)

        # Define a posição inicial para escrever os dados no PDF
        x, y = 100, 700

        # Define o espaçamento entre as linhas
        espacamento = 20

        # Define o tamanho da fonte para o nome da obra
        tamanho_fonte_nome_obra = 16

        # Adiciona o nome da obra em destaque (vermelho)
        c.setFont("Helvetica-Bold", tamanho_fonte_nome_obra)
        c.setFillColorRGB(1, 0, 0)  # Cor vermelha (RGB)
        nome_obra_text = f"Nome da Obra: {nome_obra}"
        c.drawString(x, y, nome_obra_text)
        c.setFillColorRGB(0, 0, 0)  # Retorna à cor padrão (preto)
        y -= espacamento * 2  # Afasta mais o nome da obra do nome da etapa

        # Adiciona os dados do DataFrame com uma linha de término após cada etapa
        for _, row in df.iterrows():
            etapa = f"Nome da Etapa: {row['Nome da Etapa']}"
            c.setFont("Helvetica-Bold", 12)
            c.drawString(x, y, etapa)
            y -= espacamento  # Move para a próxima linha

            responsavel = f"Responsável: {row['Responsável']}"
            c.setFont("Helvetica", 12)
            c.drawString(x, y, responsavel)
            y -= espacamento  # Move para a próxima linha

            qtd_medida = f"Qtd/Medida: {row['Qtd/Medida']}"
            c.drawString(x, y, qtd_medida)
            y -= espacamento  # Move para a próxima linha

            unidade_calculo = f"Unidade de Cálculo: {row['Unidade de Cálculo']}"
            c.drawString(x, y, unidade_calculo)
            y -= espacamento  # Move para a próxima linha

            valor_orcado = f"Valor Orçado: {row['Valor Orçado']}"
            c.drawString(x, y, valor_orcado)
            y -= espacamento  # Move para a próxima linha

            etp = f"Porcentagem da Etapa: {row['% ETP']}%"
            c.drawString(x, y, etp)
            y -= espacamento  # Move para a próxima linha

            # Adiciona uma linha de término após cada etapa
            c.line(x, y, x + 500, y)
            y -= espacamento  # Move para a próxima linha

            # Verifica se é hora de criar uma nova página
            if y < espacamento * 5:
                c.showPage()  # Cria uma nova página
                y = 700  # Reinicia a posição vertical na nova página

        # Salva o PDF
        c.save()





    def adicionar_dados_a_planilha(self):
        # Solicite ao usuário que selecione a pasta de trabalho (onde a planilha está)
        pasta_de_trabalho = filedialog.askdirectory(title="Selecione a Pasta de Trabalho")

        if not pasta_de_trabalho:
            return  # O usuário cancelou a seleção da pasta de trabalho

        # Construa o caminho completo para a planilha dentro da pasta de trabalho
        nome_da_obra = self.nomeobra_entry.get()
        caminho_planilha = os.path.join(pasta_de_trabalho, f"{nome_da_obra}_Etapas.xlsx")

        # Verifique se a planilha já existe
        if os.path.exists(caminho_planilha):
            # Carregue a planilha existente em um DataFrame do Pandas
            df_existente = pd.read_excel(caminho_planilha)
        else:
            # Se a planilha não existir, crie um DataFrame vazio
            df_existente = pd.DataFrame(columns=["Nome da Etapa", "Responsável", "Qtd/Medida", "Unidade de Cálculo",
                                                "Data de Início", "Duração (dias)", "Nome da Obra"])

        # Adicione os novos dados ao DataFrame
        nome_etapa = self.newETP_entry.get()
        responsavel = self.responseName_entry.get()
        qtd_medida = self.por_M_entry.get()
        unidade_calculo = self.comboBox.get()
        data_inicio = self.inicio_calendar_entry.get()
        duracao = self.duração_calendar_entry.get()

        novo_dado = {
            "Nome da Etapa": nome_etapa,
            "Responsável": responsavel,
            "Qtd/Medida": qtd_medida,
            "Unidade de Cálculo": unidade_calculo,
            "Data de Início": data_inicio,
            "Duração (dias)": duracao,
            "Nome da Obra": nome_da_obra
        }

        # Concatene o DataFrame existente com o novo dado e redefina o índice
        df_existente = pd.concat([df_existente, pd.DataFrame([novo_dado])], ignore_index=True)

        # Salve o DataFrame de volta na mesma planilha existente
        df_existente.to_excel(caminho_planilha, index=False)

        mensagem = f"Dados adicionados com sucesso à planilha '{os.path.basename(caminho_planilha)}'!"
        messagebox.showinfo("Sucesso", mensagem)
 

    def calcular_data_conclusao(self):
            start_date_str = self.inicio_calendar_entry.get()
            
            duration_str = self.duração_calendar_entry.get()

            try:
                start_date = datetime.strptime(start_date_str, "%d/%m/%Y")
                duration = int(duration_str)
                completion_date = start_date + timedelta(days=duration)
                self.result_label.configure(text=f"Prazo Conclusão: {completion_date.strftime('%d/%m/%Y')}",text_color="#00ff00")
            except ValueError:
                self.result_label.configure(text="Data ou Duração Invalido !",text_color="#ff0000")








sistema = TelaLogin()


