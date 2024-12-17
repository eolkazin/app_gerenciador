import tkinter as tk
from tkinter import messagebox
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill
from plyer import notification
import json
import os

class CadastroAssinante:
    def __init__(self, root):
        self.root = root
        self.root.title("Cadastro de Assinantes")

        # Variáveis para armazenar dados
        self.nome_var = tk.StringVar()
        self.telefone_var = tk.StringVar()
        self.categoria_var = tk.StringVar()
        self.valor_categoria = {"Black": 110, "Diamond": 80, "Gold": 50}

        # Contagem de assinantes por categoria
        self.contagem_categorias = {"Black": 0, "Diamond": 0, "Gold": 0}

        # Variável para rastrear o saldo
        self.saldo = 0

        # Widgets
        self.label_nome = tk.Label(root, text="Nome:")
        self.entry_nome = tk.Entry(root, textvariable=self.nome_var)

        self.label_telefone = tk.Label(root, text="Telefone:")
        self.entry_telefone = tk.Entry(root, textvariable=self.telefone_var)

        self.label_categoria = tk.Label(root, text="Categoria:")
        self.dropdown_categoria = tk.OptionMenu(root, self.categoria_var, *self.valor_categoria.keys())

        self.btn_cadastrar = tk.Button(root, text="Cadastrar", command=self.cadastrar_assinante)

        # Listbox para exibir os detalhes dos assinantes
        self.listbox_detalhes = tk.Listbox(root)
        self.listbox_detalhes.grid(row=4, column=0, columnspan=2, padx=10, pady=10, sticky=tk.W)

        # Label para exibir o saldo e a contagem de assinantes
        self.label_saldo = tk.Label(root, text="Saldo: R$0.00")
        self.label_contagem = tk.Label(root, text="Contagem de Assinantes: 0")

        # Layout
        self.label_nome.grid(row=0, column=0, padx=10, pady=10, sticky=tk.W)
        self.entry_nome.grid(row=0, column=1, padx=10, pady=10, sticky=tk.W)

        self.label_telefone.grid(row=1, column=0, padx=10, pady=10, sticky=tk.W)
        self.entry_telefone.grid(row=1, column=1, padx=10, pady=10, sticky=tk.W)

        self.label_categoria.grid(row=2, column=0, padx=10, pady=10, sticky=tk.W)
        self.dropdown_categoria.grid(row=2, column=1, padx=10, pady=10, sticky=tk.W)

        self.btn_cadastrar.grid(row=3, column=0, columnspan=2, pady=10)

        # Posiciona a Label do saldo
        self.label_saldo.grid(row=5, column=0, columnspan=2, pady=10)

        # Posiciona a Label da contagem de assinantes
        self.label_contagem.grid(row=6, column=0, columnspan=2, pady=10)

        # Carrega os dados salvos (se existirem)
        self.carregar_dados()

        # Carrega a planilha existente (se existir) ou cria uma nova
        if os.path.exists('assinantes.xlsx'):
            self.workbook = load_workbook('assinantes.xlsx')
        else:
            self.workbook = Workbook()
            sheet = self.workbook.active
            sheet.append(["Nome", "Telefone", "Categoria", "Custo"])

    def cadastrar_assinante(self):
        nome = self.nome_var.get()
        telefone = self.telefone_var.get()
        categoria = self.categoria_var.get()

        if nome and telefone and categoria:
            try:
                sheet = self.workbook.active

                # Adiciona cabeçalho se o arquivo estiver vazio
                if sheet.max_row == 1 and sheet.max_column == 1:
                    sheet.append(["Nome", "Telefone", "Categoria", "Custo"])

                # Calcula o custo com base na categoria escolhida
                custo = self.valor_categoria.get(categoria, 0)

                sheet.append([nome, telefone, categoria, custo])

                # Atualiza a contagem de assinantes por categoria
                self.contagem_categorias[categoria] += 1

                # Atualiza o saldo
                self.saldo += custo

                # Atualiza a Label do saldo
                self.label_saldo.config(text=f"Saldo: R${self.saldo:.2f}")

                # Atualiza a Label da contagem de assinantes
                total_assinantes = sum(self.contagem_categorias.values())
                self.label_contagem.config(text=f"Contagem de Assinantes: {total_assinantes}")

                # Aplica formatação condicional
                fill_color = None
                if categoria == "Black":
                    fill_color = "000000"  # Preto
                elif categoria == "Diamond":
                    fill_color = "0000FF"  # Azul
                elif categoria == "Gold":
                    fill_color = "FFD700"  # Dourado

                if fill_color:
                    cell = sheet.cell(row=sheet.max_row, column=sheet.max_column - 2)  # Aplica cor à categoria
                    cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")

                # Salva a planilha
                self.workbook.save('assinantes.xlsx')

                # Exibe uma mensagem de sucesso
                mensagem = f"Assinante cadastrado:\nNome: {nome}\nTelefone: {telefone}\nCategoria: {categoria}\nCusto: R${custo}"
                messagebox.showinfo("Cadastro realizado", mensagem)

                # Notificação
                self.mostrar_notificacao(nome, categoria, custo)

                # Atualiza o Listbox
                self.atualizar_listbox()

                # Atualiza a Label da contagem de assinantes
                total_assinantes = sum(self.contagem_categorias.values())
                self.label_contagem.config(text=f"Contagem de Assinantes: {total_assinantes}")

                # Salva os dados
                self.salvar_dados()
            except Exception as e:
                messagebox.showerror("Erro", f"Erro ao salvar a planilha: {str(e)}")
        else:
            messagebox.showerror("Erro", "Por favor, preencha todos os campos.")

    def mostrar_notificacao(self, nome, categoria, custo):
        mensagem = f"O cliente {nome} assinou a categoria {categoria} com custo R${custo}"
        notification.notify(
            title="Assinatura Realizada",
            message=mensagem,
            app_name="Cadastro de Assinantes",
        )

    def atualizar_listbox(self):
        # Limpa o Listbox
        self.listbox_detalhes.delete(0, tk.END)

        # Adiciona detalhes dos assinantes ao Listbox
        for categoria, contagem in self.contagem_categorias.items():
            self.listbox_detalhes.insert(tk.END, f"{categoria}: {contagem} assinante(s)")

    def salvar_dados(self):
        # Salva os dados em um arquivo JSON
        dados = {"contagem_categorias": self.contagem_categorias, "saldo": self.saldo}
        with open('dados.json', 'w') as arquivo:
            json.dump(dados, arquivo)

    def carregar_dados(self):
        try:
            # Tenta carregar os dados do arquivo JSON
            with open('dados.json', 'r') as arquivo:
                dados = json.load(arquivo)
                self.contagem_categorias = dados.get("contagem_categorias", {})
                self.saldo = dados.get("saldo", 0)

                # Atualiza a Label do saldo
                self.label_saldo.config(text=f"Saldo: R${self.saldo:.2f}")

                # Atualiza a Label da contagem de assinantes
                total_assinantes = sum(self.contagem_categorias.values())
                self.label_contagem.config(text=f"Contagem de Assinantes: {total_assinantes}")
        except FileNotFoundError:
            # Se o arquivo não existe, não há dados para carregar
            pass

if __name__ == "__main__":
    root = tk.Tk()
    app = CadastroAssinante(root)
    root.mainloop()
